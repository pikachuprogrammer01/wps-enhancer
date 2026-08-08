import openpyxl
from openpyxl.styles import PatternFill, Font
from typing import List

from typing import List, Optional

from core.file_io.base import (
    BaseReader, BaseWriter, SheetData, WriteRequest,
    is_declaration_first_row, is_empty_row,
)
from core.exceptions import FileReadError, FileWriteError
from core.logger import log_call


class XlsxReader(BaseReader):
    """基于 openpyxl 的 xlsx 文件读取器。"""

    @log_call("core.file_io.xlsx", log_args=True)
    def get_sheet_names(self, file_path: str) -> List[str]:
        """读取文件中所有 Sheet 名称。"""
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            return wb.sheetnames
        except Exception as e:
            raise FileReadError(f"无法读取文件 '{file_path}'：{type(e).__name__}: {e}") from e

    @log_call("core.file_io.xlsx", log_args=True)
    def read_sheet(
        self, file_path: str, sheet_name: str,
        skip_declaration: bool = False,
        declaration_keywords: Optional[List[str]] = None,
    ) -> SheetData:
        """读取指定 Sheet 的表头和数据行（可选剔除首行声明，第一行即表头）。"""
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb[sheet_name]
            all_rows = list(ws.iter_rows(values_only=True))
            wb.close()
            # 跳过前导空行（声明行前常见空行）
            while all_rows and is_empty_row(all_rows[0]):
                all_rows.pop(0)
            if not all_rows:
                raise FileReadError(
                    f"无法读取文件 '{file_path}'：Sheet '{sheet_name}' 为空"
                )
            skipped = False
            if skip_declaration and len(all_rows) >= 2 and is_declaration_first_row(
                all_rows[0], all_rows[1], declaration_keywords,
            ):
                all_rows = all_rows[1:]
                skipped = True
                # 声明行后可能跟空行
                while all_rows and is_empty_row(all_rows[0]):
                    all_rows.pop(0)
            headers = [str(cell) if cell is not None else "" for cell in all_rows[0]]
            while headers and not headers[-1]:
                headers.pop()  # 去掉尾部空表头列（文件最大列宽超出表头）
            rows = []
            for row in all_rows[1:]:
                row_dict = {}
                for i, header in enumerate(headers):
                    value = row[i] if i < len(row) else None
                    row_dict[header] = str(value) if value is not None else ""
                rows.append(row_dict)
            return SheetData(sheet_name=sheet_name, headers=headers, rows=rows, declaration_skipped=skipped)
        except FileReadError:
            raise
        except Exception as e:
            raise FileReadError(f"无法读取文件 '{file_path}'：{type(e).__name__}: {e}") from e


class XlsxWriter(BaseWriter):
    """基于 openpyxl 的 xlsx 文件写入器。"""

    @staticmethod
    def _write_headers(ws, headers: list) -> None:
        """写入加粗表头行。"""
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)

    @staticmethod
    def _write_data_rows(ws, data_rows: list) -> None:
        """写入所有数据行。"""
        for row_idx, row in enumerate(data_rows):
            for col_idx, value in enumerate(row):
                ws.cell(row=row_idx + 2, column=col_idx + 1, value=value)

    @staticmethod
    def _write_merges(ws, merge_ranges: list) -> None:
        """写入合并单元格。"""
        for merge in merge_ranges:
            start_row = merge.row_start + 2
            end_row = merge.row_end + 2
            start_col = merge.col_index + 1
            end_col = merge.col_index + 1
            ws.merge_cells(
                start_row=start_row, start_column=start_col,
                end_row=end_row, end_column=end_col,
            )
            for data_row in range(merge.row_start + 1, merge.row_end + 1):
                ws.cell(row=data_row + 2, column=merge.col_index + 1, value=None)

    @staticmethod
    def _write_cell_styles(ws, cell_styles: dict) -> None:
        """应用单元格背景色样式。"""
        for (row_idx, col_idx), style in cell_styles.items():
            if style.background_color is not None:
                color = style.background_color.lstrip("#")
                fill = PatternFill(fill_type="solid", fgColor=color)
                ws.cell(row=row_idx + 2, column=col_idx + 1).fill = fill

    @log_call("core.file_io.xlsx", log_args=True)
    def write_export(self, request: WriteRequest) -> None:
        """写入导出数据，含合并单元格和背景色标记。"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            self._write_headers(ws, request.headers)
            self._write_data_rows(ws, request.data_rows)
            self._write_merges(ws, request.merge_ranges)
            self._write_cell_styles(ws, request.cell_styles)
            wb.save(request.file_path)
        except Exception as e:
            raise FileWriteError(f"无法写入文件 '{request.file_path}'：{e}") from e
