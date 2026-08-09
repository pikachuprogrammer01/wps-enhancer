import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from core.exceptions import FileReadError, FileWriteError
from core.file_io.base import (
    SheetData, WriteRequest, CellStyle, MergeRange, is_declaration_first_row,
    get_reader, get_writer,
)


class DeclarationRowTest(unittest.TestCase):
    """企查查声明行检测（纯函数）与 reader 剔除。"""

    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp(prefix="wps_decl_"))

    def _csv(self, name: str, text: str):
        path = self.tmp / name
        path.write_text(text, encoding="utf-8")
        return path

    def test_qcc_keyword_declaration(self):
        self.assertTrue(is_declaration_first_row(["企查查"], ["姓名", "手机号"]))

    def test_multi_cell_row_never_declaration(self):
        """多格行不判声明（防误判：表头含「声明/数据来源」等词时不被跳过）。"""
        self.assertFalse(is_declaration_first_row(
            ["企查查导出数据", "2026-01-01"], ["姓名", "手机号"], ["企查查"],
        ))
        self.assertFalse(is_declaration_first_row(
            ["姓名", "数据来源"], ["张三", "企查查"], ["数据来源"],
        ))

    def test_single_cell_keyword_declaration(self):
        """单格行命中关键词即声明（即使次行单列）。"""
        self.assertTrue(is_declaration_first_row(
            ["企查查"], ["姓名"], ["企查查"],
        ))

    def test_all_empty_first_row_is_declaration(self):
        self.assertTrue(is_declaration_first_row(["", "  "], ["姓名", "手机号"]))

    def test_single_cell_first_with_multi_col_second(self):
        self.assertTrue(is_declaration_first_row(["声明"], ["姓名", "手机号"]))

    def test_normal_table_not_declaration(self):
        self.assertFalse(is_declaration_first_row(["姓名", "手机号"], ["张三", "138"]))

    def test_single_column_table_not_declaration(self):
        self.assertFalse(is_declaration_first_row(["姓名"], ["张三"]))

    def test_reader_skips_declaration_when_enabled(self):
        path = self._csv("qcc.csv", "企查查导出数据\n姓名,手机号\n张三,138\n")
        data = get_reader(str(path)).read_sheet(str(path), "qcc", skip_declaration=True)
        self.assertTrue(data.declaration_skipped)
        self.assertEqual(data.headers, ["姓名", "手机号"])
        self.assertEqual(data.rows[0]["姓名"], "张三")

    def test_reader_keeps_declaration_when_disabled(self):
        path = self._csv("qcc2.csv", "企查查导出数据\n姓名,手机号\n张三,138\n")
        data = get_reader(str(path)).read_sheet(str(path), "qcc2", skip_declaration=False)
        self.assertFalse(data.declaration_skipped)
        self.assertEqual(data.headers, ["企查查导出数据"])

    def test_declaration_with_blank_lines(self):
        """前导空行 + 单格声明行 + 声明后空行布局。"""
        path = self._csv(
            "qcc3.csv", "\n\n企查查导出数据\n\n姓名,手机号\n张三,138\n",
        )
        data = get_reader(str(path)).read_sheet(
            str(path), "qcc3", skip_declaration=True, declaration_keywords=["企查查"],
        )
        self.assertTrue(data.declaration_skipped)
        self.assertEqual(data.headers, ["姓名", "手机号"])
        self.assertEqual(len(data.rows), 1)

    def test_multi_cell_header_not_misjudged(self):
        """多格表头含关键词列名时不被误判为声明行（防误判回归）。"""
        path = self._csv(
            "mh.csv", "姓名,数据来源\n张三,企查查\n李四,天眼查\n",
        )
        data = get_reader(str(path)).read_sheet(
            str(path), "mh", skip_declaration=True,
            declaration_keywords=["企查查", "数据来源"],
        )
        self.assertFalse(data.declaration_skipped)
        self.assertEqual(data.headers, ["姓名", "数据来源"])
        self.assertEqual(len(data.rows), 2)


class CsvTest(unittest.TestCase):
    """csv 读取（编码探测）与写入（编码按设置）。"""

    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp(prefix="wps_csv_"))

    def _write(self, name: str, text: str, encoding: str = "utf-8"):
        path = self.tmp / name
        path.write_text(text, encoding=encoding)
        return path

    def test_read_utf8(self):
        path = self._write("a.csv", "姓名,手机号\n张三,138\n")
        data = get_reader(str(path)).read_sheet(str(path), "a")
        self.assertEqual(data.headers, ["姓名", "手机号"])
        self.assertEqual(data.rows[0]["手机号"], "138")

    def test_read_gbk(self):
        path = self._write("b.csv", "姓名,手机号\n李四,139\n", encoding="gbk")
        data = get_reader(str(path)).read_sheet(str(path), "b")
        self.assertEqual(data.rows[0]["姓名"], "李四")

    def test_read_bom(self):
        path = self._write("c.csv", "\ufeff姓名,手机号\n王五,137\n", encoding="utf-8")
        data = get_reader(str(path)).read_sheet(str(path), "c")
        self.assertEqual(data.headers, ["姓名", "手机号"])

    def test_empty_first_row_skipped(self):
        path = self._write("empty.csv", "\n姓名,手机号\n张三,138\n")
        data = get_reader(str(path)).read_sheet(str(path), "empty")
        self.assertEqual(data.headers, ["姓名", "手机号"])
        self.assertEqual(data.rows[0]["姓名"], "张三")

    def test_all_empty_rows_raise(self):
        path = self._write("allempty.csv", "\n\n\n")
        with self.assertRaises(FileReadError):
            get_reader(str(path)).read_sheet(str(path), "allempty")

    def test_write_bom_and_rows(self):
        path = self.tmp / "out.csv"
        request = WriteRequest(
            file_path=str(path), headers=["姓名", "手机号"],
            data_rows=[["张三", "138"], ["李四", "139"]],
            encoding="utf-8-bom",
        )
        get_writer(str(path)).write_export(request)
        raw = path.read_bytes()
        self.assertTrue(raw.startswith(b"\xef\xbb\xbf"))
        text = raw.decode("utf-8-sig")
        self.assertIn("姓名,手机号\r\n", text)
        self.assertIn("李四,139\r\n", text)


class TxtTest(unittest.TestCase):
    """txt 写入：表头行 + 自定义分隔符。"""

    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp(prefix="wps_txt_"))

    def test_write_with_separator(self):
        path = self.tmp / "out.txt"
        request = WriteRequest(
            file_path=str(path), headers=["姓名", "手机"],
            data_rows=[["张三", "138"], ["李四", "139"]],
            separator="、",
        )
        get_writer(str(path)).write_export(request)
        text = path.read_text(encoding="utf-8")
        self.assertEqual(text.splitlines()[0], "姓名、手机")
        self.assertEqual(text.splitlines()[1], "张三、138")


class VcfTest(unittest.TestCase):
    """vcf 写入：字段过滤、转义、记录结构。"""

    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp(prefix="wps_vcf_"))

    def test_write_vcf_fields_filtered(self):
        path = self.tmp / "out.vcf"
        request = WriteRequest(
            file_path=str(path),
            headers=["姓名", "手机", "公司名", "网址"],
            data_rows=[["张三", "13800000000", "某某公司", "https://a.com"]],
            field_keys=["name", "phone", "company", "website"],
            vcf_fields=["name", "phone", "company"],
        )
        get_writer(str(path)).write_export(request)
        text = path.read_text(encoding="utf-8")
        self.assertIn("BEGIN:VCARD", text)
        self.assertIn("FN:张三", text)
        self.assertIn("TEL;TYPE=CELL:13800000000", text)
        self.assertIn("ORG:某某公司", text)

    def test_write_vcf_name_prefix_suffix(self):
        """vcf 姓名前后缀生效（FN 带前缀后缀，便于通讯录批量管理）。"""
        path = self.tmp / "prefix.vcf"
        request = WriteRequest(
            file_path=str(path),
            headers=["姓名", "手机"],
            data_rows=[["张三", "13800000000"]],
            field_keys=["name", "phone"],
            vcf_fields=["name", "phone"],
            vcf_name_prefix="客户-",
            vcf_name_suffix="-VIP",
        )
        get_writer(str(path)).write_export(request)
        text = path.read_text(encoding="utf-8")
        self.assertIn("FN:客户-张三-VIP", text)
        self.assertIn("TEL;TYPE=CELL:13800000000", text)
        self.assertNotIn("URL:", text)  # website 不在 vcf_fields
        self.assertIn("END:VCARD", text)

    def test_write_escapes_special_chars(self):
        path = self.tmp / "esc.vcf"
        request = WriteRequest(
            file_path=str(path),
            headers=["姓名", "手机"],
            data_rows=[["张,三;四", "138"]],
            field_keys=["name", "phone"],
        )
        get_writer(str(path)).write_export(request)
        text = path.read_text(encoding="utf-8")
        self.assertIn("FN:张\\,三\\;四", text)

    def test_write_escapes_carriage_return(self):
        path = self.tmp / "cr.vcf"
        request = WriteRequest(
            file_path=str(path),
            headers=["姓名", "手机"],
            data_rows=[["张\r三", "138"]],
            field_keys=["name", "phone"],
        )
        get_writer(str(path)).write_export(request)
        text = path.read_text(encoding="utf-8")
        self.assertIn("FN:张\\r三", text)

    def test_write_folds_long_lines(self):
        path = self.tmp / "fold.vcf"
        long_url = "https://example.com/" + "x" * 100
        request = WriteRequest(
            file_path=str(path),
            headers=["姓名", "手机", "网址"],
            data_rows=[["张三", "138", long_url]],
            field_keys=["name", "phone", "website"],
        )
        get_writer(str(path)).write_export(request)
        raw = path.read_bytes()
        self.assertIn(b"\r\n ", raw)  # vCard 3.0 折叠续行（原始字节）
        for line in raw.split(b"\r\n"):
            if line.startswith(b"URL:"):
                self.assertLessEqual(len(line), 75)

    def test_missing_field_keys_raises(self):
        path = self.tmp / "bad.vcf"
        request = WriteRequest(file_path=str(path), headers=["姓名"], data_rows=[["张三"]])
        with self.assertRaises(FileWriteError):
            get_writer(str(path)).write_export(request)

    def test_empty_values_skipped(self):
        path = self.tmp / "empty.vcf"
        request = WriteRequest(
            file_path=str(path),
            headers=["姓名", "手机"],
            data_rows=[["张三", ""]],
            field_keys=["name", "phone"],
        )
        get_writer(str(path)).write_export(request)
        text = path.read_text(encoding="utf-8")
        self.assertNotIn("TEL", text)


try:
    import openpyxl  # noqa: F401
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


try:
    import xlrd  # noqa: F401
    import xlwt  # noqa: F401
    HAS_XL = True
except ImportError:
    HAS_XL = False


@unittest.skipUnless(HAS_OPENPYXL, "openpyxl 未安装")
class XlsxRoundTripTest(unittest.TestCase):
    """xlsx 写入→读取回环。"""

    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp(prefix="wps_xlsx_"))

    def test_round_trip_with_style_and_merge(self):
        path = self.tmp / "out.xlsx"
        request = WriteRequest(
            file_path=str(path),
            headers=["姓名", "手机"],
            data_rows=[["张三", "138"], ["", "139"]],
            merge_ranges=[MergeRange(row_start=0, row_end=1, col_index=0)],
            cell_styles={(1, 1): CellStyle(background_color="#FF0000")},
        )
        get_writer(str(path)).write_export(request)
        data = get_reader(str(path)).read_sheet(str(path), "Sheet")
        self.assertEqual(data.headers, ["姓名", "手机"])
        self.assertEqual(data.rows[0]["手机"], "138")

    def test_header_is_first_row_even_when_data_row_has_more_cells(self):
        """表头行非空单元格少于数据行时，表头仍取第一行（回归：旧启发式误判）。"""
        path = self.tmp / "hdr.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["姓名", "手机号"])          # 表头：2 个非空
        ws.append(["张三", "138", "多余列"])   # 数据行：3 个非空（旧逻辑会误选为表头）
        ws.append(["李四", "139"])
        wb.save(path)
        data = get_reader(str(path)).read_sheet(str(path), "Sheet")
        self.assertEqual(data.headers, ["姓名", "手机号"])
        self.assertEqual(len(data.rows), 2)
        self.assertEqual(data.rows[0]["姓名"], "张三")
        # 企查查开关开启但无声明行时同样取第一行
        data2 = get_reader(str(path)).read_sheet(str(path), "Sheet", skip_declaration=True)
        self.assertEqual(data2.headers, ["姓名", "手机号"])
        self.assertFalse(data2.declaration_skipped)

    def test_merged_declaration_cell_detected(self):
        """第一行合并单元格声明（A1:F1 合并，仅左上角有字）可被识别跳过。"""
        path = self.tmp / "merged_decl.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.merge_cells("A1:F1")
        ws["A1"] = "企查查导出数据"
        ws.append(["姓名", "手机号", "公司", "网址", "备注", "地区"])
        ws.append(["张三", "138", "A公司", "http://a.com", "无", "北京"])
        wb.save(path)
        data = get_reader(str(path)).read_sheet(str(path), "Sheet", skip_declaration=True)
        self.assertTrue(data.declaration_skipped)
        self.assertEqual(data.headers, ["姓名", "手机号", "公司", "网址", "备注", "地区"])
        self.assertEqual(len(data.rows), 1)
        self.assertEqual(data.rows[0]["姓名"], "张三")


@unittest.skipUnless(HAS_XL, "xlrd/xlwt 未安装")
class XlsRoundTripTest(unittest.TestCase):
    """xls 写入→读取回环。"""

    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp(prefix="wps_xls_"))

    def test_round_trip(self):
        path = self.tmp / "out.xls"
        request = WriteRequest(
            file_path=str(path),
            headers=["姓名", "手机"],
            data_rows=[["张三", "138"], ["李四", "139"]],
        )
        get_writer(str(path)).write_export(request)
        data = get_reader(str(path)).read_sheet(str(path), "Sheet1")
        self.assertEqual(data.headers, ["姓名", "手机"])
        self.assertEqual(data.rows[1]["手机"], "139")

    def test_header_is_first_row_even_when_data_row_has_more_cells(self):
        """表头行非空单元格少于数据行时，表头仍取第一行（回归：旧启发式误判）。"""
        path = self.tmp / "hdr.xls"
        wb = xlwt.Workbook()
        ws = wb.add_sheet("Sheet1")
        ws.write(0, 0, "姓名")
        ws.write(0, 1, "手机号")          # 表头：2 个非空
        ws.write(1, 0, "张三")
        ws.write(1, 1, "138")
        ws.write(1, 2, "多余列")          # 数据行：3 个非空
        ws.write(2, 0, "李四")
        ws.write(2, 1, "139")
        wb.save(str(path))
        data = get_reader(str(path)).read_sheet(str(path), "Sheet1")
        self.assertEqual(data.headers, ["姓名", "手机号"])
        self.assertEqual(len(data.rows), 2)
        self.assertEqual(data.rows[0]["姓名"], "张三")


class UnsupportedFormatTest(unittest.TestCase):
    """不支持的格式抛异常。"""

    def test_unsupported_reader(self):
        with self.assertRaises(FileReadError):
            get_reader("/tmp/x.ods")

    def test_unsupported_writer(self):
        with self.assertRaises(FileWriteError):
            get_writer("/tmp/x.ods")


if __name__ == "__main__":
    unittest.main()


class SheetSummariesTest(unittest.TestCase):
    """get_sheet_summaries：多 sheet 名称+行数（下拉区分纯数字 sheet）。"""

    def test_xlsx_multi_sheet_summaries(self):
        import openpyxl
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "multi.xlsx"
            wb = openpyxl.Workbook()
            ws1 = wb.active
            ws1.title = "1"
            ws1.append(["姓名", "手机"])
            ws1.append(["张三", "13800000000"])
            ws2 = wb.create_sheet("联系人")
            ws2.append(["姓名"])
            ws2.append(["李四"])
            ws2.append(["王五"])
            wb.save(path)
            summaries = get_reader(str(path)).get_sheet_summaries(str(path))
        self.assertEqual(summaries, [("1", 2), ("联系人", 3)])

    def test_csv_single_sheet_summary(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "data.csv"
            path.write_text("姓名,手机\n张三,13800000000\n", encoding="utf-8")
            summaries = get_reader(str(path)).get_sheet_summaries(str(path))
        self.assertEqual(len(summaries), 1)
        self.assertEqual(summaries[0][0], "data")
        self.assertGreaterEqual(summaries[0][1], 2)
